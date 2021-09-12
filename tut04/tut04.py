import os
import csv
import openpyxl
from openpyxl import Workbook
os.system("cls")


"""  TASK-1:
     Function to make a file for every individual ROLL NUMBER 
    listed in the "regtable old.csv". All the outputs relating
     to Task 1 will go into "output_individual_roll" folder.
"""

def output_individual_roll():

    """Creating folder output_individual_roll for storing outputs
       for task 1.
    """
    os.mkdir("./output_individual_roll")

    """loop_no2 to check if it's the first iteration in the
        file , we skip the first iteration or first line
        of the file containing variable names/labels
    """

    loop_no2=0

    """reading the file regtable_old.csv"""

    with open('regtable_old.csv', 'r') as fr:
        reader=csv.reader(fr)
        for records in reader:

            """if first iteration then skip it"""

            loop_no2+=1
            if loop_no2==1:
                continue

            """if the the file does not exists then create file 
               and add headers as the first line
            """

            headers=[]
            headers.append("rollno")
            headers.append("register_sem")
            headers.append("subno")
            headers.append("sub_type")

            details=[]
            details.append(records[0])
            details.append(records[1])
            details.append(records[3])
            details.append(records[-1])


            if os.path.isfile("./output_individual_roll/"+records[0]+".xlsx")==False:
                book = Workbook()
                sheet = book.active
                sheet.append(headers)
                book.save("./output_individual_roll/"+records[0]+".xlsx")

            """adding the data i.e, the details related to the roll number"""
            book = openpyxl.load_workbook("./output_individual_roll/"+records[0]+".xlsx")
            sheet = book.active
            sheet.append(details)
            book.save("./output_individual_roll/"+records[0]+".xlsx")
        fr.close()

    return



"""  TASK-2:
     Function to make a file for every individual subject 
    listed in the "regtable old.csv". All the outputs relating
     to Task 2 will go into "output_by_subject" folder.
"""

def output_by_subject():

    """Creating folder output_by_subject for storing outputs
       for task 2.
    """
    os.mkdir("./output_by_subject")

    """loop_no to check if it's the first iteration in the
        file , we skip the first iteration or first line
        of the file containing variable names/labels
    """
    loop_no=0

    """reading the file regtable_old.csv"""

    with open('regtable_old.csv', 'r') as f:
        reader2=csv.reader(f)
        for record in reader2:

            """if first iteration then skip it"""

            loop_no+=1
            if loop_no==1:
                continue

            """if the the file does not exists then create file 
               and add headers as the first line
            """

            headers=[]
            headers.append("rollno")
            headers.append("register_sem")
            headers.append("subno")
            headers.append("sub_type")

            details=[]
            details.append(record[0])
            details.append(record[1])
            details.append(record[3])
            details.append(record[-1])


            if os.path.isfile("./output_by_subject/"+record[3]+".xlsx")==False:
                book = Workbook()
                sheet = book.active
                sheet.append(headers)
                book.save("./output_by_subject/"+record[3]+".xlsx")

            """adding the data i.e, the details related to the subject"""
            book = openpyxl.load_workbook("./output_by_subject/"+record[3]+".xlsx")
            sheet = book.active
            sheet.append(details)
            book.save("./output_by_subject/"+record[3]+".xlsx")
        f.close()
    return
   
# TASK : 1
output_individual_roll()
# TASK : 2
output_by_subject()