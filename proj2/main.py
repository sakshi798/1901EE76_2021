# importing all the required libraraies or modules
import csv    
import os
import re
import openpyxl  
import pandas as pd      
from fpdf import FPDF
from datetime import date
from datetime import datetime                      
from openpyxl.workbook import Workbook
from pathlib import Path

path = str(Path.home()) + "\\Desktop\\Proj2_Python"
# Dictionary to store overall result of all semesters for a rollno
rollno_to_overall={}
list_rollno=[]
# Dictionary that will be used to access pointer from grade(integer)
grade_dict = {'AA' : 10, 'AB' : 9, 'BB': 8, 'BC': 7,'CC' : 6,  'CD' : 5, 'DD' : 4,'DD*' : 4, 'F' : 0, 'F*' : 0, ' BB' : 8}

# Dictionary for converting month number to month name for date of Issue
month_num= {1:"Jan",2:"Feb",3:"March",4:"April",5:"May",6:"June",7:"July",8:"Aug",9:"Sep",10:"Oct", 11:"Nov", 12:"Dec"}

# Column Names for Semester Wise sheets
headings = ['Sub Code','Subject Name','L-T-P','CRD','GRD']

# Dictionary that will be used to access name from roll number
Dict_Name_RollNo = {}

def output(path1,path2,path3):
        
    # Read file and fill list (rollno) and dictionaries (roll to name)
    file_pointer_nr = open(path2, 'r')
    File_Name_RollNumber = csv.DictReader(file_pointer_nr)

    for x in File_Name_RollNumber:
        Dict_Name_RollNo[x['Roll']] = x['Name']
        list_rollno.append(str(x['Roll']).upper())

    # Dictionary that will be used to access subject name,ltp,credits from subject number
    Dict_Subj_Master = {}

    # Read file and fill dictionaries (roll to name,ltp,credits from subject number)
    file_pointer_sm = open(path3, 'r')
    File_Subj_Master = csv.DictReader(file_pointer_sm)
    for x in File_Subj_Master:
        Dict_Subj_Master[x['subno']] = [x['subname'],x['ltp'],x['crd']]

    # Dictionary that with keys roll number and semester, so we can access details
    # corresponding to each roll number and semester
    Dicti_Grades = {}
    # Read file and fill dictionaries
    file_grades = open(path1, 'r')
    File_Grades = csv.DictReader(file_grades)
    for x in File_Grades:
    # if roll number is not present then add a dictionary
    # with the key as rollnumber and add info
        if not x['Roll'] in Dicti_Grades.keys():
            Dicti_Grades[x['Roll']] = {}
            # if semester is present then append the info
            # in the key as rollnumber and semester
            if x['Sem'] in Dicti_Grades[x['Roll']].keys():
                Dicti_Grades[x['Roll']][x['Sem']].append([ x['SubCode'],Dict_Subj_Master[x['SubCode']][0],Dict_Subj_Master[x['SubCode']][1],
                                          Dict_Subj_Master[x['SubCode']][2],x['Grade']])
            else:
                Dicti_Grades[x['Roll']][x['Sem']] = [[ x['SubCode'],Dict_Subj_Master[x['SubCode']][0],Dict_Subj_Master[x['SubCode']][1],
                                          Dict_Subj_Master[x['SubCode']][2],x['Grade']]]
        # if roll number is present then add the info
        # to the dict with key as roll no   
        else:
            # if semester is present then append the info
            # in the key as rollnumber and semester
            if x['Sem'] in Dicti_Grades[x['Roll']].keys():
                Dicti_Grades[x['Roll']][x['Sem']].append([ x['SubCode'], Dict_Subj_Master[x['SubCode']][0],Dict_Subj_Master[x['SubCode']][1],
                                            Dict_Subj_Master[x['SubCode']][2],x['Grade']])
            # if semester is not present then intialise info
            # in the key as rollnumber and semester  
            else:
                Dicti_Grades[x['Roll']][x['Sem']] = [[ x['SubCode'],Dict_Subj_Master[x['SubCode']][0],Dict_Subj_Master[x['SubCode']][1],
                                          Dict_Subj_Master[x['SubCode']][2], x['Grade']]]
    # Loop for each roll number in the dictionary 
    for x in Dicti_Grades.keys():
        New_Workbook = Workbook()

        Sem_Wise_Credits = ['Semester wise Credit Taken']
        Semester_Number = ['Semester No.']
        
        # Make output directory if it is not already present
        Name_Of_File = path + "\\output\\" + x + ".xlsx"
        Folder_Output = os.path.dirname(Name_Of_File)
        if not os.path.exists(Folder_Output):
            os.makedirs(Folder_Output)
        Total_Credits_Taken = ['Total Credits Taken']
        SPI = ['SPI']
        CPI = ['CPI'] 
        # Adding Rollno, name from dictionary, and Discpline to the overall sheet
        overall = []
        overall.append(['Roll No.', x])
        overall.append(['Name of the Student',Dict_Name_RollNo[x]])
        overall.append(['Discipline',x[4:6]])  
        # Calculation for all semesters
        for Semester in Dicti_Grades[x].keys():
            Semester_Number.append(Semester)
            Serial_Number = 1 
    
            # Create separate sheets for all semesters
            # Name of the sheet must be Sem (Semester_number) 
            New_Worksheet = New_Workbook.create_sheet()
            New_Worksheet.title = 'Sem' + Semester
            New_Worksheet.append(headings)

            creditss = 0
            spii = 0 
    # Calculating spi for the semester using grade pointer and credits
    # SPI = (C1 ∗ G1 + C2 ∗ G2 + C3 ∗ G3 + C4 ∗ G4)/(C1 + C2 + C3 + C4)
            for Details in Dicti_Grades[x][Semester]:
                spii += int(Details[3]) * grade_dict[Details[4]]
                creditss += int(Details[3]) 
                New_Worksheet.append(Details)
                Serial_Number += 1
            Sem_Wise_Credits.append(creditss)
            spi=spii/creditss
            # Adding spi 
            SPI.append(round(spi,2))
            # Calculating cpi for the semester using spi and credits
            # CPI = (SPI1 ∗ Credits in semester1 + SPI2 ∗ Credits in semester2 + ...)/(Total credits) 
            if type(Total_Credits_Taken[-1]) == str:
                Total_Credits_Taken.append(creditss)
                cpi=spii/creditss
                CPI.append(round(cpi,2))
            else:
                Total_Cred = Total_Credits_Taken[-1] + creditss
                calc_spi=round((CPI[-1]*Total_Credits_Taken[-1]+ spii)/Total_Cred,2)
                CPI.append(calc_spi)
                Total_Credits_Taken.append(Total_Cred)
        # Creating new worksheet for overall details
        New_Worksheet = New_Workbook['Sheet'] 
        New_Worksheet.title='Overall'
        # Adding semester number, credits, SPI,credits and CPI
        overall.append(Semester_Number),overall.append(Sem_Wise_Credits),overall.append(SPI)
        overall.append(Total_Credits_Taken),overall.append(CPI)
        
        rollno_to_overall[x]=[]
        rollno_to_overall[x].append(overall)
        for x in overall:
            New_Worksheet.append(x)
        # Save the changes in the workbook
        New_Workbook.save(filename=Name_Of_File)

# Function to create pdf for a particular rollno
def pdf_for_rollno(rollno,path4,path5):
    
    # Calculating date and time of generation
    todays_date = date.today()
    cur_yr = int(str(todays_date.year)[-2:])
    if int(rollno[:2]) <= cur_yr:
	    year = str(cur_yr-1) + str(rollno[:2]) 
    elif int(rollno[:2]) > cur_yr:
	    year = str(cur_yr-2) + str(rollno[:2])
    today=date.today()
    d1=today.strftime("%d.%m.%Y")
    dated=str(d1[0:2])+" "
    month_num= {1:"Jan",2:"Feb",3:"March",4:"April",5:"May",6:"June",7:"July",8:"Aug",9:"Sep",10:"Oct", 11:"Nov", 12:"Dec"}
    dated = dated + month_num[int(d1[3:5])]+" "
    dated+=str(d1[6:11])
    current_time = datetime.now().strftime("%H:%M:%S")

    # Find name for rollno using dictionary
    name=Dict_Name_RollNo[rollno]
    
    # Find Programme name using rollno
    if rollno[2:4] == "01":
	    Prog = " Bachelor of Technology"	
    elif rollno[2:4] == "11":
        Prog = " Master of Technology"
    elif rollno[2:4] == "12":
        Prog = " Master of Science"
    elif rollno[2:4] == "21":
        Prog = " Doctor of Philosophy"

    # Find overall details for the rollno and semester
    credits_taken=rollno_to_overall[rollno][0][4]
    credits_cleared=rollno_to_overall[rollno][0][5]
    SPI=rollno_to_overall[rollno][0][6]
    CPI=rollno_to_overall[rollno][0][7]

    # Class with function definition for pdf 
    class PDF(FPDF):

        # Function to add rectangles in pdf
        def lines(self):

            # Border of pdf using rectangle
            self.rect(10.0, 10.0,399,275.5)
            # Box for name, rollno, year, etc details
            self.rect(84.0, 44 ,250.0, 13.0) 
            #Textbox for roll no
            self.rect(105, 44.5 ,28.0, 4.5)
            #Textbox for Name
            self.rect(170, 44.5 ,83.0, 4.5)
            #Textbox for year
            self.rect(316, 44.5 ,15.0, 4.5)
            
        # Function to add text in pdf
        def titles(self):

            # Add text "INTERIM TRANSCRIPT" below first logo
            self.set_xy(23.0,23.3)
            self.set_font('Arial', 'U', 7)
            self.set_text_color(0, 0, 0)
            self.cell(w=20.0, h=30.0, align='C', txt="INTERIM TRANSCRIPT", border=0)
            
            # Add text "INTERIM TRANSCRIPT" below second logo
            self.set_xy(376.5,23.3)
            self.set_font('Arial', 'U', 7)
            self.set_text_color(0, 0, 0)
            self.cell(w=20.0, h=30.0, align='C', txt="INTERIM TRANSCRIPT", border=0)
            
            # Add text "Roll No:" below  IITP heading
            self.set_xy(80.5,32.0)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=50.0, h=30.0, align='C', txt="Roll No:        "+rollno, border=0)
            
            # Add text "Programme:" below  IITP heading
            self.set_xy(87.5,38.0)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=20.0, h=30.0, align='C', txt="Programme:", border=0)
            
            # Add text programme name below  IITP heading
            self.set_xy(114.5,38.0)
            self.set_font('Arial', '', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=20.0, h=30.0, align='C', txt=Prog, border=0)
            
            # Add text "Name:" below  IITP heading
            self.set_xy(150.5,32.0)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=64.0, h=30.0, align='C', txt="Name:             "+name, border=0)
            
            # Add text "Course:" below  IITP heading
            self.set_xy(148.5,38.0)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=34.0, h=30.0, align='C', txt="Course:    ", border=0)
            
            # Add text Course name below  IITP heading
            self.set_xy(190.5,38.0)
            self.set_font('Arial', '', 9.5)
            self.set_text_color(0, 0, 0)
            et = [x.upper() for x in rollno if not (ord(x) >=48 and ord(x)<=57)]
            et = "".join(et)
            
            # Course assigned to roll number on the basis 
            # of the branch extracted from roll number.
            if et=="CS":
                crs = "Computer Science and Engineering"
            elif et=="ME":
                crs = "Mechanical Engineering"
            elif et=="EE":
                crs = "Electrical Engineering"
            elif et=="CH":
                crs = "Chemical and Biochemical Engineering"
            elif et=="MME":
                crs = "Metallurgical and Materials Engineering"
            elif et=="CE":
                crs = "Civil Engineering"
            self.cell(w=20.0, h=30.0, align='C', txt=crs, border=0)

            # Add text "Year of Admission:" and year below  IITP heading
            self.set_xy(289.5,32.0)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=30.0, h=30.0, align='C', txt="Year of Admission:        "+year, border=0)
            
            # Add text "Assistant Registrar (Academic)" below  IITP heading
            self.set_xy(363.5,242.5)
            self.set_font('Arial', 'B', 9.5)
            self.set_text_color(0, 0, 0)
            self.cell(w=20.0, h=10.0, align='C', txt="Assistant Registrar (Academic)", border=0)
            
        # Function to add images and lines int the pdf
        def imagex(self,x):

            # Add Line below the iitp heading
            self.set_line_width(0.0) 
            self.line(10,40,409,40) 

            # Add standing line beside logo 1
            self.set_line_width(0.0) 
            self.line(59,10,59,39.7)

            # Add standing line beside logo 2  
            self.set_line_width(0.0) 
            self.line(360,10,360,39.7)

            # Add IITP logo at top-left corner 
            self.set_xy(6.0,6.0)
            self.image(name = str(Path.home()) + "\\Desktop\\logo_short.jpg", x=11,y=12, type='JPG', w=3500/80, h=2100/80)
            # Add IITP heading at top-mid part
            self.set_xy(12.0,16.0)
            self.image(name = str(Path.home()) + "\\Desktop\\iitp.png", x=61 ,y=14, type='PNG', w=295, h=2000/80)  
            # Add IITP logo at top-right corner 
            self.set_xy(18.0,16.0)
            self.image(name = str(Path.home()) + "\\Desktop\\logo_short.jpg", x=363,y=12, type='JPG', w=3500/80, h=2100/80)
            
            y2=x+(280-x)/2
            # Add the date of issue line bottom left
            self.set_line_width(0.0) 
            self.line(40,y2,70,y2) 

            # Add signature image if it is uploaded and path exists
            if os.path.exists(path5):
                self.set_xy(340.0,y2-28)#stamp
                self.image(name = path5, x=342 ,y=y2-28, type='PNG', w=35, h=35)
            else:
                pass 

            #Add the signature line bottom right of pdf
            self.set_line_width(0.0) 
            self.line(348,y2+4,400,y2+4) 

            # Add stamp image if it is uploaded and path exists
            if os.path.exists(path4):
                self.set_xy(112.0,y2-28)#stamp
                self.image(name = path4, x=172 ,y=y2-28, type='PNG', w=50, h=50)  
            
        # Function to add tables (marksheets) and overall result  
        def marksheets(self):

            # Line below the heading
            self.set_line_width(0.0) 

            # Variable to calculate max y(height) of the table
            # so that we can calculate position of other table and line

            y21max=0    # end of first row of Tables 
            y22max=0    # end of second row of Tables 

            # Read excel sheets ( already created) to print marksheet 
            wb = openpyxl.load_workbook(path + "\\output\\" + rollno + ".xlsx") 

            # Number of sheets in the workbook
            res = len(wb.sheetnames)

            # Traverse all semester sheets in workbook
            for j in range(1,res):
                
                # Postion calculation
                x1=12 + 99.5*((j-1)%4)
                if j<5:
                    y1=56
                else:
                    y1=y21max+10

                # Add semester number heading
                self.set_xy(x1,y1)
                self.set_font('Arial', 'BU', 9.5)
                self.set_text_color(0, 0, 0)
                self.cell(w=20.0, h=10.0, align='C', txt="Semester "+str(j), border=0)

                # Calculation of position
                self.set_font('arial', 'BU', 7)
                self.x=12 + 99.5*((j-1)%4)
                if j<5:
                    self.y=63
                else:
                    self.y=y21max+30

                # Read the excel sheets and print marksheet table
                df = pd.read_excel(path + "\\output\\" + rollno + ".xlsx",sheet_name=j)
                col_width = [17,57,7,7,7]
                columnNameList = list(df.columns)
                
                # number of rows in table
                n=(len(df.index))+1
                i=0
                self.set_font('Arial', 'B', 7)

                # Traverse in column names of table
                for header in columnNameList:

                    if i!=len(columnNameList)-1:
                        self.cell(col_width[i], 5, header, 1, 0, 'C')
                    else :
                        break
                    i+=1
                
                self.cell(col_width[i], 5, columnNameList[-1], 1, 2, 'C')
                self.cell(-88)
                self.set_font('arial', '', 7)

                # Write tables in pdf
                for row in range(0, len(df)):
                    i1=-1
                    for col_num, col_name in enumerate(columnNameList):
                        i1+=1
                        if col_num != len(columnNameList)-1:
                            self.cell(col_width[i1],5, str(df['%s' % (col_name)].iloc[row]), 1, 0, 'C')
                        else:
                            self.set_font('Arial', 'B', 7)
                            self.cell(col_width[i1],5, str(df['%s' % (col_name)].iloc[row]), 1, 2, 'C')  
                            self.cell(-88)
                            self.set_font('Arial', '', 7)
                
                # Calculation of position
                x2=12 + 99.5*((j-1)%4)
                if j<5:
                    y2=65+n*5
                    y21max=max(y21max,y2)
                    y22max=y21max
                else:
                    y2=y21max+32+n*5
                    y22max=max(y22max,y2)

                # Add heading and overall details like credits taken, credits cleared
                # SPI , CPI etc below the table
                det=" "
                det=det+" Credits Taken:   "+str(credits_taken[j])
                det+=" "
                det=det+" Credits Cleared:   "+str(credits_cleared[j])
                det+=" "
                det=det+" SPI:   "+str(SPI[j])
                det+=" "
                det=det+" CPI:   "+str(CPI[j])

                #Textbox details overall per sem  
                self.rect(x2, y2 ,78.0, 5.3)#Textbox details overall per sem  
                self.set_xy(x2+8,y2-12)
                self.set_font('Arial', 'B', 7)
                self.set_text_color(0, 0, 0)
                self.cell(w=60.0, h=30.0, align='C', txt=det, border=0)

            # Add line 1 after table 1
            self.set_line_width(0.0)
            self.line(10,y21max+10,409,y21max+10) 
            # Add line  after table 
            self.set_line_width(0.0) 
            self.line(10,y22max+18,409,y22max+18) 

            return (y22max+18)  
                          
                    
    # Create pdf , A3 sheet landscape mode 
    pdf = PDF(orientation='L', unit='mm', format='A3')

    # ADD PAGE
    pdf.add_page()

    # Calll functions to add image and text
    pdf.lines()
    pdf.imagex(pdf.marksheets())
    pdf.titles()

    # Add date of issue left bottom
    pdf.set_xy(23.5,232.0)
    pdf.set_font('Arial', 'B', 9.5)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(w=35.0, h=10.0, align='C', txt="Date of Issue:    "+dated+" "+str(current_time[0:5]), border=0)

    # save changes with given file name     
    pdf.output(path + "\\transcriptsIITP\\" + rollno.upper() + ".pdf",'F')
    return

# Function to generate pdf for a given range
def generate_pdf_range(rmin,rmax,path4,path5):

    # If the folder does not exists then create one
    if not os.path.isdir(path + "\\transcriptsIITP"):
        os.mkdir(path + "\\transcriptsIITP")

    # Change to uppercase letters
    Rmin=rmin.upper()
    Rmax=rmax.upper()

    # if the function calls for all rollnos then generate all pdfs
    if (Rmin==list_rollno[0]) and (Rmax==list_rollno[len(list_rollno)-1]):
        for i in range(0,len(list_rollno)):
            pdf_for_rollno(list_rollno[i],path4,path5)
        return

    # Else check if the given range is valid and then
    # Generate pdf for valid roll numbers
    else :
        
        # Interpret the roll number range
        roll_list = [Rmin,Rmax]
        roll_range=Rmin+"-"+Rmax

        # Use regex expression matching to check if the 
        # range of roll numbers entered by user is correct
        roll_range_exp = r"(([0-9]+[01,11,12,13]+[a-zA-Z]+[0-9]+)(?:-)([0-9]+[01,11,12,13]+[a-zA-Z]+[0-9]+))"
        
        # Check if the roll no matches the pattern and is valid  
        if bool(re.match(roll_range_exp, roll_range)):

            # Check if both the roll numbers are of same branch
            if (roll_list[0][4:6].upper() != roll_list[1][4:6].upper()) or (int(roll_list[0][6:]) > int(roll_list[1][6:])):
                return r"Enter valid range of roll number"

        # Pattern does not match then range is not valid
        else:
            return r"Enter valid range of roll numbers"

        # Pop the list of roll numbers which don't exist and
        # generate transcripts for those roll numbers that exist 
        rollno_start = roll_list[0][6:].lstrip("0")
        rollno_end = roll_list[1][6:].lstrip("0")

        # Calculate start and end in integers
        low_roll = int(rollno_start)
        high_roll = int(rollno_end) + 1

        # List to store roll numbers that don't exist
        roll_not_exist = []

        # Traverse through all the numbers
        for i in range(low_roll, high_roll):

            rn = roll_list[0][:6]
            st_rn = ""

            for ch in rn:
                # Check using ASCII value if the character is in lower case or upper case
                if (ord(ch)>=65 and ord(ch)<=90) or (ord(ch)>=97 and ord(ch)<=122):
                    st_rn+=ch.upper()
                else:
                    st_rn+=ch

            # Calculate roll no as string format year-branch-rollno
            if len(str(i))<2:
                roll_num = st_rn + "0"*(2-len(str(i))) +str(i)
            else:
                roll_num = st_rn + str(i)

            # Count number of occurences of roll no
            exist_count = list_rollno.count(roll_num)
 
            # checking if it is more then 0
            if exist_count <= 0:
                roll_not_exist.append(roll_num)
            else:
                pdf_for_rollno(roll_num,path4,path5)

        return roll_not_exist

# Function to generate all pdfs
def generate_pdf_all(path4,path5):

    # Call function to generate in range
    # range = (first rollno,last rollno) from rollno list
    generate_pdf_range(list_rollno[0],list_rollno[len(list_rollno)-1],path4,path5)

