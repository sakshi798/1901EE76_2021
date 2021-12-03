import os 
import csv
import pandas as pd
from shutil import copyfile
from openpyxl.styles import Font
from csv import writer
from openpyxl import Workbook,drawing                                          #importing various necessary modules from openpyxl to get
from openpyxl.styles import Alignment,Border,Side                              #the required marksheet designs
import smtplib
from email.message import EmailMessage
from pathlib import Path

EMAIL_ADDRESS = "mailbotcs384@gmail.com"                                       #defining the email address and password of the email ID
EMAIL_PASSWORD = "mailbotCS384PythonProject1"                                  #that will be used in sending emails to all students

font1=Font(bold=True,underline="single",size=18,name="Century")                #creating different font styles to apply on different text
font2=Font(size=12,name="Century")
font3=Font(bold=True,size=12,name="Century")

def assign_cell(ws,cellno,values,fonts):                                       #defining a function to assign value and font
    cell=ws[cellno]
    cell.value=values
    cell.font=fonts

def assign_with_border(ws,cellno,values,fonts,bor):                            #defining a function to assign value, border and font
    cell=ws[cellno]
    cell.value=values
    cell.font=fonts
    cell.border = bor

def create_concise_mk(path1,path2,pos,neg):                                    #defining a function to create concise marksheet

    pos = float(pos)                                                           #obtaining the positive and negative marks from form
    neg = float(neg)

    path = str(Path.home()) + "\\Desktop\\Proj1_Python\\marksheets"            #creating a path for marksheets (output) folder

    with open(path1, mode='r') as inp1:
        reader = csv.reader(inp1)
        rollno_to_name = {rows[0]:rows[1] for rows in reader}                  #creating a dictionary to store name for every roll number
    with open(path2, mode='r') as inp2:
        reader = csv.reader(inp2)
        rollno_to_details = {rows[6]:rows for rows in reader}                  #creating a dictionary to store responses rows for roll number
    dst = path
    try: 
        os.mkdir(dst)                                                          #creating marksheets folder if it doesn't exist already
    except OSError as error: 
        pass

    src = path2
    dst = path + "\\concise_marksheet.csv"                                     #defining the path for concise marksheet inside marksheets folder
    List_final_marks=[]
    statusAns=[]
    Absent_list=[]

    copyfile(src, dst)                                                         #creating a copy of responses in dst from src

    for rollno in rollno_to_name:                                              #traversing through all the roll numbers in dictionary
        if rollno != "roll":
            if rollno in rollno_to_details:
                this_no=[]
                right=0                                                            #number of right, wrong and not answered questions
                not_ans=0
                wrong=0

                marking_right = pos                                                #marking Scheme for right, wrong and unanswered
                marking_wrong = neg
                not_ans_mark=0

                list1 = rollno_to_details[rollno]                                  #storing responses for every roll number
                ans_list = rollno_to_details["ANSWER"]                             #storing correct responses (answer key) 

                for i in range(7,len(list1)):                                      #checking the responses for all questions
                    if ans_list[i]==list1[i]:                                      #calculating the number of right/wrong/not answered
                        right=right+1                                              #questions and calculating marks using these
                    elif list1[i]=="":
                        not_ans=not_ans+1
                    else:
                        wrong=wrong+1

                if rollno in rollno_to_details:                                    #creating a list for final marks

                    this_no.append(right)
                    this_no.append(wrong)
                    this_no.append(not_ans)
                    statusAns.append(this_no)                                      #calculating the final marks and adding to the list
                    List_final_marks.append(str(marking_right*right+marking_wrong*wrong+not_ans_mark*not_ans)+"/"+str((len(list1)-7)*marking_right))
            
            else:
                List=['','','ABSENT']
                List.append(rollno_to_name[rollno])
                List.append("")
                List.append("")
                List.append("ABSENT")
                List.append(str(rollno))
                
                Absent_list.append(List)
 
    df = pd.read_csv(dst)                                                      #using pandas to rename and add column names according
    new_df = df.rename(columns = {"Score": "Google_Score"})                    #to output in copied file responses
    new_df.insert(6, "Score_After_Negative", List_final_marks)
    new_df.insert(36,"statusAns",statusAns)
    new_df.to_csv(dst,index=False)

    for lists in Absent_list:
        with open(dst, 'a') as f_object:
                    writer_object = writer(f_object)
                    writer_object.writerow(lists)
                    f_object.close() 


def create_rollnowise(path1,path2,pos,neg):                                    #defining a function to create roll no wise marksheets

    pos = float(pos)                                                           #obtaining the positive and negative marks from form
    neg = float(neg)

    path = str(Path.home()) + "\\Desktop\\Proj1_Python\\marksheets"            #creating a path for marksheets (output) folder 

    with open(path1, mode='r') as inp1:
        reader = csv.reader(inp1)
        rollno_to_name = {rows[0]:rows[1] for rows in reader}                  #creating a dictionary to store name for every roll number
    with open(path2, mode='r') as inp2:
        reader = csv.reader(inp2)
        rollno_to_details = {rows[6]:rows for rows in reader}                  #creating a dictionary to store responses rows for roll number
    dst = path
    try: 
        os.mkdir(dst)                                                          #creating marksheets folder if it doesn't exist already
    except OSError as error: 
        pass

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),    #defining border style for cells
                    top=Side(style='thin'), bottom=Side(style='thin'))

    src = path2
    dst = path + '\\concise_marksheet.csv'                                    #defining the path for concise marksheet file
    img_path = str(Path.home()) + '\\Desktop\\IITP_Logo.jpg'                  #defining the path for image on desktop

    ans_list = rollno_to_details["ANSWER"]                                    #answer key for checking answers of students

    for rollno in rollno_to_name:                                             #traversing through all roll numbers
        if rollno != "roll" and rollno in rollno_to_details:
            name=rollno+".xlsx"                                               #defining name for excel sheet
            wb = Workbook()
            ws1 = wb.create_sheet("quiz")                                     #creating workbook for every roll number
            
            img = drawing.image.Image(img_path)                               #inserting IITP logo in marksheet
            img.width = 619
            img.height = 76
            img.anchor = 'A1'
            ws1.add_image(img)

            assign_cell(ws1,"C5","Mark sheet",font1)                          #adding cells in excel sheet
            assign_cell(ws1,"A6","Name: ",font2)
            assign_cell(ws1,"A7","Roll Numer: ",font2)
            assign_cell(ws1,"D6","Exam: ",font2)
            
            assign_cell(ws1,"E6","quiz",font3)
            assign_cell(ws1,"B6",rollno_to_name[rollno],font3)
            assign_cell(ws1,"B7",rollno,font3)

            assign_with_border(ws1,"B9","Right",font3,thin_border)
            assign_with_border(ws1,"C9","Wrong",font3,thin_border)
            assign_with_border(ws1,"D9","Not Attempt",font3,thin_border)
            assign_with_border(ws1,"E9","Max",font3,thin_border)

            assign_with_border(ws1,"A10","No.",font2,thin_border)
            assign_with_border(ws1,"A11","Marking",font2,thin_border)
            assign_with_border(ws1,"A12","Total",font2,thin_border)

            assign_with_border(ws1,"A15","Student Ans",font3,thin_border)
            assign_with_border(ws1,"B15","Correct Ans",font3,thin_border)
            assign_with_border(ws1,"D15","Student Ans",font3,thin_border)
            assign_with_border(ws1,"E15","Correct Ans",font3,thin_border)

            right=0                                                            #defining number of right, wrong and unanswered questions
            not_ans=0
            wrong=0

            marking_right = pos                                                #marking Scheme of right, wrong and unanswered questions
            marking_wrong = neg
            not_ans_mark=0

            cell=ws1["B11"]                                                    #adding cells in excel sheet
            cell.value=marking_right
            cell.font=Font(color="008000",size=12,name="Century")
            cell.border = thin_border

            cell=ws1["C11"]                                                    #adding cells in excel sheet
            cell.value=marking_wrong
            cell.font=Font(color="FF0000",size=12,name="Century")
            cell.border = thin_border

            assign_with_border(ws1,"D11",not_ans_mark,font2,thin_border)       #adding border to the cell

            list1 = rollno_to_details[rollno]                                  #declaring a list to store details for a roll number

            assign_with_border(ws1,"E10",str(len(list1)-7),font2,thin_border)  #adding border to the cell

            for i in range(7,len(list1)):                                      #checking all answers and calculating marks using marking scheme
                r=16+i-7                                                       #calculating cell name
                if(r<41):
                    s="A"+str(r)
                else:
                    s="D"+str(r-41+16)

                if(r<41):
                    s2="B"+str(r)
                else:
                    s2="E"+str(r-41+16)

                cell=ws1[s2]                                                   #assigning value,font and border to cell
                cell.value=ans_list[i]
                cell.font=Font(color="0000FF",size=12,name="Century")
                cell.border = thin_border

                cell=ws1[s]                                                    #assigning value to the cell
                cell.value=list1[i]

                if ans_list[i]==list1[i]:                                      #changing color according to whether the response is correct or not
                    cell.font=Font(color="008000",size=12,name="Century")
                    cell.border = thin_border
                else:
                    cell.font=Font(color="FF0000",size=12,name="Century")
                    cell.border = thin_border

                if ans_list[i]==list1[i]:                                      #counting wrong and right 
                    right=right+1
                elif list1[i]=="":
                    not_ans=not_ans+1
                else:
                    wrong=wrong+1

            cell=ws1["B10"]                                                    #assigning value,font and border to cell
            cell.value=(right)
            cell.font=Font(color="008000",size=12,name="Century")
            cell.border = thin_border

            cell=ws1["C10"]
            cell.value=(wrong)
            cell.font=Font(color="FF0000",size=12,name="Century")
            cell.border = thin_border

            cell=ws1["D10"]
            cell.value=(not_ans)
            cell.font=Font(size=12,name="Century")
            cell.border = thin_border

            cell=ws1["B12"]
            cell.value=marking_right*right
            cell.font=Font(color="008000",size=12,name="Century")
            cell.border = thin_border

            cell=ws1["C12"]
            cell.value=marking_wrong*wrong
            cell.font=Font(color="FF0000",size=12,name="Century")
            cell.border = thin_border

            cell=ws1["D12"]
            if not_ans_mark*not_ans!=0:
                cell.value=not_ans_mark*not_ans
                cell.font=Font(size=12,name="Century")
            cell.border = thin_border

            cell=ws1["E12"]
            cell.value=str(marking_right*right+marking_wrong*wrong+not_ans_mark*not_ans)+"/"+str((len(list1)-7)*marking_right)
            
            cell.font=Font(color="0000FF",size=12,name="Century")
            cell.border = thin_border

            for col in ws1.columns:                                                                     #changing the alignment for all cells
                for cell in col:
                    # openpyxl styles aren't mutable,
                    # so we have to create a copy of the style, modify the copy, then set it back
                    cell.alignment = Alignment(horizontal="center",vertical="bottom")

            dims = {}
            for row in ws1.rows:                                                                        #adjusting width of the cells
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws1.column_dimensions[col].width = value 
            cols=['A','B','C','D','E']
            for col in cols:
                ws1.column_dimensions[col].width = 17.09
            wb.remove(wb['Sheet'])                                                                      #removing the extra sheet created
            wb.save(path + '\\' + name)                                                                 #saving the changes in the marksheet folder

def send_email():                                                                                       #defining a function to send emails
    dir = str(Path.home()) + "\\Desktop\\Proj1_Python\\marksheets"                                      #obtaining the path of the marksheets folder
    path = dir + "\\concise_marksheet.csv"

    for file in os.listdir(dir):                                                                        #iterating through all the files present in marksheets folder
        filelist = file.split('.')                                                                      #obtaining the name of a file without extension
        nomaillist = ['concise_marksheet','ANSWER']

        if filelist[0] not in nomaillist:                                                               #making sure that we don't send an email to ANSWER
            with open(path,'r') as f:                                                                   #opening concise_marksheet.csv for every student's marksheet
                reader = csv.reader(f)
                for row in reader:
                    if row[7] == filelist[0]:                                                           #for each student we obtain their 2 email IDs and store them
                        contacts = [row[1],row[4]]                                                      #in contacts list

                        msg = EmailMessage()
                        msg['Subject'] = "CS384 2021 Python Quiz Marks with Negative"                   #setting up the basic parts of an email
                        msg['From'] = EMAIL_ADDRESS
                        msg['To'] = contacts                                                            #sending the email with attachment to both IDs
                        msg.set_content("CS384 2021 Quiz marks are attached for reference.")

                        pth = dir + "\\" + file                                                         #obtaining the path of the concerned file

                        with open(pth,'rb') as f:                                                       #adding the concerned xlsx file as an attachment
                            file_data = f.read()
                            file_name = file
                            msg.add_attachment(file_data,maintype = "application",subtype = "xlsx",filename = file_name)

                        with smtplib.SMTP_SSL('smtp.gmail.com',465) as smtp:                            #establishing a port for SMTP to operate
                            smtp.login(EMAIL_ADDRESS,EMAIL_PASSWORD)                                    #using SSL for encryption
                            smtp.send_message(msg)


